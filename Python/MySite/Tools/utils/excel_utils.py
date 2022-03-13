from openpyxl import Workbook
import openpyxl
import string_utils
import pandas as pd
import os
import glob
import xlsxwriter


def write_to_excel(arr, filename):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'data'
    for row in range(1, len(arr)+1):
        for col in range(1, len(arr[0])+1):
            sheet.cell(column=col, row=row,
                       value="{0}".format(arr[row-1][col-1]))

    workbook.save(filename=filename)
    workbook.close()
    

def read_from_excel(path): 
    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    
    # Get workbook active sheet object
    # from the active attribute
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
        
    lst = []
    
    # Loop will print all columns name
    for i in range(1, max_row + 1):
        arr = []
        for j in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row = i, column = j)     
            if(cell_obj.value is not None):
                print(cell_obj.value)   
                arr.append(string_utils.compound_unicode(str(cell_obj.value)))
            else:
                arr.append('')
                
        lst.append(arr)
            
    return lst
    

def change_column():
    # use glob to get all the csv files
    # in the folder
    path = r'C:\Users\ADMIN\Downloads\New folder (2)'
    csv_files = glob.glob(os.path.join(path, "*.xlsx"))

    # loop over the list of excel files
    for f in csv_files:

        # read the excel file
        df = pd.read_excel(f)

        df_new = pd.read_excel(f)

        # set header
        df_new.columns = ['Số tt', 'Sổ KHVB', 'Ngày tháng VB',
                        'Trích yếu nội dung', 'Tác giả văn bản', 'Tờ số', 'ghi chú']

        df_new['Số tt'] = df['Số tt']
        df_new['Sổ KHVB'] = df['Sổ KHVB']
        df_new['Ngày tháng VB'] = df['Ngày tháng VB']
        df_new['Trích yếu nội dung'] = df['Trích yếu nội dung']
        df_new['Tác giả văn bản'] = df['Tác giả văn bản']
        df_new['Tờ số'] = df['Tờ số']
        df_new['ghi chú'] = df['ghi chú']

        os.remove(f)
        writer = pd.ExcelWriter(f, engine='xlsxwriter')

        # df_new.to_excel(f, index=False)

        df_new.to_excel(writer, sheet_name='Sheet1', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        border_fmt = workbook.add_format(
            {'bottom': 2, 'top': 2, 'left': 2, 'right': 2})
        worksheet.conditional_format(xlsxwriter.utility.xl_range(0, 0, len(df_new), len(
            df_new.columns) - 1), {'type': 'no_errors', 'format': border_fmt})

        format_header = workbook.add_format()
        format_header.set_valign('vcenter')
        format_header.set_align('center')
        format_header.set_bold()
        format_header.set_text_wrap()

        format_data = workbook.add_format()
        format_data.set_valign('vcenter')
        format_data.set_align('center')
        format_data.set_text_wrap()

        worksheet.set_column('A:Z', 25, format_data)
        worksheet.set_row(0, 30, format_header)

        writer.save()

change_column()