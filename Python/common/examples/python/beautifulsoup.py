from bs4 import BeautifulSoup
from openpyxl import Workbook

# Reading the data inside the xml
# file to a variable under the name
# data
with open(r'C:\M-Files Content Packages\Kho của tôi\Out\{F20982B5-ED09-4625-8DEE-7500D35D42C2}_20221018_084502\Metadata\Content1.xml', 'r', encoding="utf-8") as f:
    data = f.read()

# Passing the stored data inside
# the beautifulsoup parser, storing
# the returned object
Bs_data = BeautifulSoup(data, "xml")


# Using find() to extract attributes
# of the first instance of the tag
# properties = Bs_data.find_all('property', {'name': 'Tên hoặc Tiêu đề'})
property = Bs_data.find_all('property')

properties = Bs_data.find_all('properties')

workbook = Workbook()
sheet = workbook.active
sheet.title = 'data'
# sheet.column_dimensions['A'].width = 50
# sheet.column_dimensions['B'].width = 100

col = 1
row = 1
dict = {}

# append to dict
for pro in property:
    # print("{0:50}\t{1}".format(property['name'], property.text))
    a = pro['name']
    if pro['name'] not in dict:
        dict[pro['name']] = [pro.text]
    else:
        dict[pro['name']].append(pro.text)

# print title
for key, value in dict.items():
    sheet.cell(column=col, row=1,
               value="{0}".format(key))
    col += 1


# print value
col = 1
row = 1
for values in dict.values():
    for value in values:
        row += 1
        sheet.cell(column = col, row = row, value="{0}".format(value))
    row = 1
    col += 1


# save excel
workbook.save(filename=r'C:\Users\Nam\Downloads\M-Files\data.xlsx')
workbook.close()
