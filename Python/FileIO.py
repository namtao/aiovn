import os
# import datetime
# import math
# import pyodbc
# import pandas as pd
# import json
# from pathlib import Path
from openpyxl import Workbook
# import numpy as np
# from openpyxl.utils import get_column_letter
# import os, shutil
# from anytree import Node, RenderTree

# get file in format
def getFiles (folderPath, txtPath, fileFormat):
    with(open(txtPath, "w", encoding="utf-8")) as f:
        lst = []
        count = 0
        # root: print path folder from file
        # dirs: sub-directories from root
        # files: path files
        for root, dirs, files in os.walk(folderPath):
            for file in files:
                if file.endswith("." + fileFormat):
                    count +=1
                    # path file
                    # print(os.path.join(root, file))
                    # folderName = Path(os.path.join(root, file)).parent.name
                    # # test folder name
                    # if not folderName[:3].isdigit() and not folderName[-3:].isdigit():
                    #     lst.append(root)
        
        # Remove Duplicates from list
        # for i in list(dict.fromkeys(lst)):
        #     f.write(i + '\n')
        # print('Có ' + str(len(list(dict.fromkeys(lst)))) + ' không mã')
        # f.close()
        f.write('Có ' + str(count) + ' file jpg')

# def datetimeFormat():
#     x = datetime.datetime.today().date()
#     print(x.strftime("%d/%m/%Y"))

# def getDB():
#     # create connection string db
#     conn = pyodbc.connect('Driver={SQL Server};'
#                       'Server=.;'
#                       'Database=ADDJ_AnGiang;'
#                       'Trusted_Connection=yes;')

#     # print console
#     cursor = conn.cursor()
#     cursor.execute('select metadata from TblMetadata')
#     # print((list(cursor)))

#     i = 0
#     for row in cursor:
#         # print(row[0])
#         # print(json.dumps(row[0], ensure_ascii=False, indent = 1))
#         data = json.loads(row[0])
#         # print(data)
#         for key in data:
#             for x in key:
#                 if x == 'indexName': indexName = key.get(x)
#                 if x == 'indexValue': indexValue = key.get(x)
#                 if x == 'indexValue2': indexValue2 = key.get(x)
#                 if x == 'indexValueQC': indexValueQC = key.get(x)

#             if len(indexValueQC.strip()) > 0: i += 1

#             # if indexValue != indexValue2: print("1 khác 2")
#             # print(key.get(x))
#     print('Có ' + str(i) + ' trường')
    
#     # convert cursor to array
#     cursor = conn.cursor()
#     cursor.execute('select metadata from TblMetadata')
#     arr = np.array(list(cursor))
#     writeToExcel(arr)

def writeToExcel(arr):
    
    filename = r"C:\Projects\Python\test.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'data'
    for row in range(1, len(arr)+1):
        for col in range(1, len(arr[0])+1):
            sheet.cell(column=col, row=row, value="{0}".format(arr[row-1][col-1]))

    # for row in range(1,101):
    #     for col in range(1,101):
    #         sheet.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))

    workbook.save(filename=filename)
    workbook.close()

# def copytree(src, dst, symlinks=False, ignore=None):    
#     if not os.path.exists(dst):  os.mkdir(dst)
#     for item in os.listdir(src):
#         s = os.path.join(src, item)
#         d = os.path.join(dst, item)
#         if os.path.isdir(s):
#             shutil.copytree(s, d, symlinks, ignore)
#         else:
#             shutil.copy2(s, d)

# def createtree():
    # udo = Node("Udo")
    # marc = Node("Marc", parent=udo)
    # lian = Node("Lian", parent=marc)
    # dan = Node("Dan", parent=udo)
    # jet = Node("Jet", parent=dan)
    # jan = Node("Jan", parent=dan)
    # joe = Node("Joe", parent=dan)
    # joe2 = Node("nam", parent=joe)

    # print(RenderTree(udo).by_attr())

if __name__ == "__main__":
    # getFiles(os.getcwd(), r'\\192.168.31.206\Share\JPG (chưa kiểm tra)\quét', 'jpg')
    # getFiles(os.path.dirname(__file__), os.path.dirname(__file__) + r'\count.txt', 'jpg')
    print('a')