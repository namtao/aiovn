from numpy.lib import utils
from openpyxl import Workbook
import openpyxl
import numpy as np
from string_utils import compound_unicode


def write_to_excel(arr, filename):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'data'
    for row in range(1, len(arr)+1):
        for col in range(1, len(arr[0])+1):
            sheet.cell(column=col, row=row,
                       value="{0}".format(arr[row-1][col-1]))

    workbook.save(filename=filename)
    workbook.close()
    

def read_from_excel(path): 
    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    
    # Get workbook active sheet object
    # from the active attribute
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
        
    lst = []
    
    # Loop will print all columns name
    for i in range(1, max_row + 1):
        arr = []
        for j in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row = i, column = j)     
            if(cell_obj.value is not None):
                print(cell_obj.value)   
                arr.append(compound_unicode(str(cell_obj.value)))
            else:
                arr.append('')
                
        lst.append(arr)
            
    return lst
    

