import difflib
from datetime import datetime

import openpyxl
import pandas as pd
from openpyxl import Workbook


def is_date(timestamp_string):
    format_string = "%Y-%m-%d %H:%M:%S"

    # checking if format matches the date
    res = True

    # using try-except to check for truth value
    try:
        res = bool(datetime.strptime(timestamp_string, format_string))
    except ValueError:
        res = False

    return res


def read_from_excel(path):
    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)

    # Get workbook active sheet object
    # from the active attribute
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row

    # list all row excel
    lst = []

    # list title
    lstTitle = []

    # Loop will print all rows name
    for i in range(1, max_row + 1):
        # list noi dung
        arr = []

        # Loop will print all column name
        for j in range(1, max_col + 1):
            # doc thong tin trong o
            cell_obj = sheet_obj.cell(row=i, column=j)

            if cell_obj.value is not None:
                # compound unicode cell_obj
                content = str(cell_obj.value)

                # nếu là dòng đầu tiên thì thêm vào list tiêu đề
                # if i == 1:
                #     lstTitle.append(content)
                # else:
                #     arr.append(content)
                if is_date(content):
                    content = (
                        datetime.strptime(content, "%Y-%m-%d %H:%M:%S")
                        .date()
                        .strftime("%d/%m/%Y")
                    )

                arr.append(content)

            else:
                # nếu trống thông tin list noi dung se them ''
                arr.append("")

        # thêm vào danh sách nội dung dòng
        lst.append(arr)

    return lst


def write_to_excel(arr, filename):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "data"
    for row in range(1, len(arr) + 1):
        for col in range(1, len(arr[0]) + 1):
            sheet.cell(column=col, row=row, value="{0}".format(arr[row - 1][col - 1]))

    workbook.save(filename=filename)
    workbook.close()


# Thay đổi định dạng file
write_to_excel(
    read_from_excel("/home/vannam/Downloads/0.xlsx"), "/home/vannam/Downloads/3.xlsx"
)


goc0 = pd.read_excel("/home/vannam/Downloads/0.xlsx")
goc = pd.read_excel("/home/vannam/Downloads/3.xlsx", header=None)


COL_ID = {
    1: "Hộp số HS",
    2: "Hồ sơ số HSS",
    3: "Tiêu đề hồ sơ hoặc trích yếu nội dung",
    4: "Số tờ Tờ số",
    5: "Ngày tháng BĐ-KT Ngày tháng văn bản",
    6: "Năm",
    7: "Thời hạn bảo quản THBQ",
    8: "Ghi chú",
    9: "Số ký hiệu VB",
    10: "Cơ quan ban hành Tác giả văn bản",
    11: "Loại tài liệu",
    12: "Bộ phận",
    13: "MLHS",
    14: "STT",
    15: "Tên loại văn bản",
    24: "Số tạm",
}

lstVitri = []
for i in list(goc0.columns.ravel()):
    val = difflib.get_close_matches(i, COL_ID.values(), cutoff=0.5)
    if val is None:
        val = difflib.get_close_matches(i, COL_ID.values(), cutoff=0.3)

    index = None
    for k, v in COL_ID.items():
        if len(val) > 0 and val[0] == v:
            lstVitri.append(k)
            index = k
            break

    COL_ID.pop(index)


vitri = []

vitri.insert(0, lstVitri)

# Thêm cột vị trí
goc = pd.concat([pd.DataFrame(vitri), goc], ignore_index=True)
goc.to_excel("/home/vannam/Downloads/3.xlsx", index=False, header=False)
