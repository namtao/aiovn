import configparser
import datetime
import os
from pathlib import Path
import re

import numpy as np
import openpyxl
import pandas as pd
from sqlalchemy.engine import create_engine
from urllib.parse import quote_plus
from collections import Counter, OrderedDict
import xlsxwriter
from openpyxl.styles import Border, Side

config = configparser.ConfigParser()
config.read(r"projects\config.ini", encoding="utf-8")
sections = config.sections()
lst_huyen = sections[1:]
tk_path = config["global"]["tk"]
ngaythang = f"{datetime.date.today()}"

ks = config["global"]["tkks"]
kt = config["global"]["tkkt"]
kh = config["global"]["tkkh"]
cmc = config["global"]["tkcmc"]
hn = config["global"]["tkhn"]

ksfmt = config["global"]["ksfmt"]
ktfmt = config["global"]["ktfmt"]
khfmt = config["global"]["khfmt"]
cmcfmt = config["global"]["cmcfmt"]
hnfmt = config["global"]["hnfmt"]

sql_details_by_status = config["global"]["sql_details_by_status"]
sql_details_by_book = config["global"]["sql_details_by_book"]
sql_details_by_year = config["global"]["sql_details_by_year"]

dic_loai = {
    ks: "HT_KHAISINH",
    kt: "HT_KHAITU",
    kh: "HT_KETHON",
    cmc: "HT_NHANCHAMECON",
    hn: "HT_XACNHANHONNHAN",
}


def get_database_connection(config, section: str):
    connection_string = f"mssql://{config['global']['user']}:{quote_plus(config['global']['pass'])}@{config['global']['host']}/{config[section]['db']}?driver={config['global']['driver']}"
    return create_engine(connection_string)


def read_sql_query(conn, sql) -> pd.DataFrame:
    df = pd.read_sql_query(sql, conn)
    df.replace(r"^\s*$", np.nan, regex=True, inplace=True)
    return df


def tktruong(conn, sql):
    return np.sum(read_sql_query(conn, sql).count())


def removeEscape(value):
    return " ".join(str(value).splitlines()).strip()


def count_non_empty_cells(df):
    return np.sum(df.count())


def merge_dict(dict1, dict2):
    # for i in dict2.keys():
    #     if i in dict1:
    #         dict1[i] = dict1[i] + dict2[i]
    #     else:
    #         dict1[i] = dict2[i]
    return {k: dict1.get(k, 0) + dict2.get(k, 0) for k in set(dict1) | set(dict2)}


def reduce_mem_usage(df):
    """iterate through all the columns of a dataframe and modify the data type
    to reduce memory usage.
    """
    start_mem = df.memory_usage().sum() / 1024**2
    print(f"Memory usage of dataframe is {start_mem:.2f} MB")

    for col in df.columns:
        col_type = df[col].dtype

        if (
            col_type != object
            and col_type.name != "category"
            and "datetime" not in col_type.name
        ):
            c_min = df[col].min()
            c_max = df[col].max()
            if str(col_type)[:3] == "int":
                if c_min > np.iinfo(np.int8).min and c_max < np.iinfo(np.int8).max:
                    df[col] = df[col].astype(np.int8)
                elif c_min > np.iinfo(np.int16).min and c_max < np.iinfo(np.int16).max:
                    df[col] = df[col].astype(np.int16)
                elif c_min > np.iinfo(np.int32).min and c_max < np.iinfo(np.int32).max:
                    df[col] = df[col].astype(np.int32)
                elif c_min > np.iinfo(np.int64).min and c_max < np.iinfo(np.int64).max:
                    df[col] = df[col].astype(np.int64)
            else:
                if (
                    c_min > np.finfo(np.float16).min
                    and c_max < np.finfo(np.float16).max
                ):
                    df[col] = df[col].astype(np.float16)
                elif (
                    c_min > np.finfo(np.float32).min
                    and c_max < np.finfo(np.float32).max
                ):
                    df[col] = df[col].astype(np.float32)
                else:
                    df[col] = df[col].astype(np.float64)
        elif "datetime" not in col_type.name:
            df[col] = df[col].astype("category")

    end_mem = df.memory_usage().sum() / 1024**2
    print(f"Memory usage after optimization is: {end_mem:.2f} MB")
    print(f"Decreased by {100 * (start_mem - end_mem) / start_mem:.1f}%")

    return df


def tk_so_luong(conn, sql):
    df = pd.read_sql_query(sql, conn)
    return int(removeEscape(df.to_string(index=False)))


def liet_ke_ndk(conn, sql):
    df = pd.read_sql_query(sql, conn)
    return df.to_dict()


def ten_noi_dang_ky(conn, sql):
    df = pd.read_sql_query(sql, conn)
    return df.to_dict()


def tk_so_truong_by_df(df, so_ky_tu_duoi: int, so_ky_tu_tren: int):
    # tạo dataframe từ câu lênh sql
    # thực thi sql
    # df = pd.read_sql_query(sql, conn)
    # print(len(np.nan))
    for col in df.columns:
        df[col] = df[col].apply(
            lambda x: x if so_ky_tu_duoi <= len(str(x)) <= so_ky_tu_tren else " "
        )
    df.replace(r"^\s*$", np.nan, regex=True, inplace=True)
    return count_non_empty_cells(df)  # đếm số ô có thông tin (loại bỏ nan)


def export_excel(file_name, df):
    writer = pd.ExcelWriter(file_name, engine="xlsxwriter")
    # pd.DataFrame(dic_truong).to_excel(
    #     writer, sheet_name="Thống kê hộ tịch", index=False
    # )

    pd.DataFrame(df).to_excel(writer, sheet_name="Thống kê hộ tịch", index=False)

    workbook = writer.book
    worksheet = writer.sheets["Thống kê hộ tịch"]

    format_number = workbook.add_format({"num_format": "#,##0"})

    format_percentage = workbook.add_format({"num_format": "00.00%"})

    # Set the column width and format.

    format_data = workbook.add_format()
    format_data.set_valign("vcenter")
    # format_data.set_align('center')
    format_data.set_text_wrap()

    worksheet.set_column("A:Z", 10, format_number)
    writer.close()


def export_excel_format(loai_so, folder_path, huyen, sql, conn):
    export_excel(
        os.path.join(
            folder_path,
            f"Danh sách sai định dạng - {loai_so}.xlsx",
        ),
        pd.read_sql_query(sql, conn),
    )


def tk_chi_tiet_theo_nam(
    conn,
    ks_sql,
    kt_sql,
    kh_sql,
    cmc_sql,
    hn_sql,
    file_name_ban_ghi,
    file_name_truong,
    ndk=0,
    chi_tiet=True,
):
    dic_truong = {
        "Năm": [],
        "KS 1 15": [],
        "KS 16 50": [],
        "KS 51": [],
        "KT 1 15": [],
        "KT 16 50": [],
        "KT 51": [],
        "KH 1 15": [],
        "KH 16 50": [],
        "KH 51": [],
        "CMC 1 15": [],
        "CMC 16 50": [],
        "CMC 51": [],
        "HN 1 15": [],
        "HN 16 50": [],
        "HN 51": [],
    }

    dic_ban_ghi = {
        "Năm": [],
        "KS": [],
        "KT": [],
        "KH": [],
        "CMC": [],
        "HN": [],
    }

    df_nam = pd.DataFrame([])

    df_ks = pd.read_sql(ks_sql, conn)
    df_kt = pd.read_sql(kt_sql, conn)
    df_kh = pd.read_sql(kh_sql, conn)
    df_cmc = pd.read_sql(cmc_sql, conn)
    df_hn = pd.read_sql(hn_sql, conn)

    if not chi_tiet:
        for sql_data, table_name in dic_loai.items():
            df_nam = pd.concat(
                [
                    df_nam,
                    pd.DataFrame(
                        (
                            pd.read_sql(sql_data, conn)["quyenSo"].str[-4:].copy()
                        ).drop_duplicates()
                    ),
                ]
            )
    else:
        df_nam = pd.concat(
            [
                df_nam,
                pd.DataFrame(
                    (
                        pd.read_sql(f"{ks} where noiDangKy = {ndk}", conn)["quyenSo"]
                        .str[-4:]
                        .copy()
                    ).drop_duplicates()
                ),
                pd.DataFrame(
                    (
                        pd.read_sql(f"{kt} where noiDangKy = {ndk}", conn)["quyenSo"]
                        .str[-4:]
                        .copy()
                    ).drop_duplicates()
                ),
                pd.DataFrame(
                    (
                        pd.read_sql(f"{kh} where noiDangKy = {ndk}", conn)["quyenSo"]
                        .str[-4:]
                        .copy()
                    ).drop_duplicates()
                ),
                pd.DataFrame(
                    (
                        pd.read_sql(f"{cmc} where noiDangKy = {ndk}", conn)["quyenSo"]
                        .str[-4:]
                        .copy()
                    ).drop_duplicates()
                ),
                pd.DataFrame(
                    (
                        pd.read_sql(f"{hn} where noiCap = {ndk}", conn)["quyenSo"]
                        .str[-4:]
                        .copy()
                    ).drop_duplicates()
                ),
            ]
        )

    df_nam = df_nam.drop_duplicates().sort_values("quyenSo")
    df_nam.reset_index(inplace=True, drop=True)
    df_nam = pd.DataFrame(df_nam)

    # df_ks["quyenSo"] = df_ks["quyenSo"].str[-4:]
    # grouped = df_ks.groupby('quyenSo')

    # b = df_ks.groupby(['quyenSo']).size().reset_index(name = 'SL')
    # print(b)

    for nam in df_nam["quyenSo"].to_list():
        print(nam)
        dic_truong["Năm"].append(nam)
        dic_ban_ghi["Năm"].append(nam)

        lst_key = ["Năm"]
        lst_key_ban_ghi = ["Năm"]

        # thong ke truong thong tin theo nam
        for loai in [df_ks, df_kt, df_kh, df_cmc, df_hn]:
            lst_index = [1, 15, 16, 50, 51, 10000]
            index = 0
            for i in dic_truong.keys():
                if not i in lst_key and index < 5:
                    lst_key.append(i)

                    df_child = pd.DataFrame([])
                    df_child = loai.loc[loai["quyenSo"].str.contains(f"/{nam}")].copy()

                    dic_truong[i].append(
                        tk_so_truong_by_df(
                            df_child,
                            lst_index[index],
                            lst_index[index + 1],
                        )
                    )
                    index += 2

        # thong ke so luong ban ghi theo nam
        for loai in [df_ks, df_kt, df_kh, df_cmc, df_hn]:
            for i in dic_ban_ghi.keys():
                if not i in lst_key_ban_ghi:
                    df_child = pd.DataFrame([])
                    df_child = loai.loc[loai["quyenSo"].str.contains(f"/{nam}")].copy()

                    dic_ban_ghi[i].append(df_child.shape[0])
                    lst_key_ban_ghi.append(i)
                    break

    # print(dic_truong)
    # print(dic_ban_ghi)

    export_excel(file_name_ban_ghi, dic_ban_ghi)
    export_excel(file_name_truong, dic_truong)
    pretty_excel(file_name_truong)


def tk_chi_tiet_theo_nam_xa(
    fmt=True, status=True, book=True, year=True, period=True, detail=True, general=True
):
    # for ma_cap_cha, huyen in dic.items():
    for huyen in lst_huyen:
        ma_cap_cha = config[huyen]["mcc"]
        conn = get_database_connection(config, huyen)
        print(huyen)

        # Xuat danh sach loi
        if fmt:
            new_path = os.path.join(tk_path, ngaythang, "Lỗi định dạng", huyen)
            Path(new_path).mkdir(parents=True, exist_ok=True)

            export_excel_format("KS", new_path, huyen, ksfmt, conn)
            export_excel_format("KT", new_path, huyen, ktfmt, conn)
            export_excel_format("KH", new_path, huyen, khfmt, conn)
            export_excel_format("CMC", new_path, huyen, cmcfmt, conn)
            export_excel_format("HN", new_path, huyen, hnfmt, conn)

        # Xuat danh sach thong ke theo trang thai
        if status:
            new_path = os.path.join(
                tk_path,
                ngaythang,
                "Thống kê chi tiết theo trạng thái",
            )
            Path(new_path).mkdir(parents=True, exist_ok=True)

            export_excel(
                os.path.join(
                    new_path,
                    f"{huyen}.xlsx",
                ),
                pd.read_sql_query(sql_details_by_status, conn),
            )
        # Xuat danh sach thong ke theo quyen so
        if book:
            new_path = os.path.join(
                tk_path,
                ngaythang,
                "Thống kê số lượng theo quyển số",
            )
            Path(new_path).mkdir(parents=True, exist_ok=True)

            export_excel(
                os.path.join(
                    new_path,
                    f"{huyen}.xlsx",
                ),
                pd.read_sql_query(sql_details_by_book, conn),
            )

        # Xuat danh sach thong ke theo nam
        if year:
            new_path = os.path.join(
                tk_path,
                ngaythang,
                "Thống kê số lượng theo năm mở sổ",
            )
            Path(new_path).mkdir(parents=True, exist_ok=True)

            export_excel(
                os.path.join(
                    new_path,
                    f"{huyen}.xlsx",
                ),
                pd.read_sql_query(sql_details_by_book, conn),
            )

        # Xuat danh sach thong ke theo khoang thoi gian
        if period:
            new_path = os.path.join(
                tk_path,
                ngaythang,
                "Thống kê số lượng theo khoảng thời gian",
            )
            Path(new_path).mkdir(parents=True, exist_ok=True)

            export_excel(
                os.path.join(
                    new_path,
                    f"{huyen}.xlsx",
                ),
                pd.read_sql_query(sql_details_by_book, conn),
            )

        # Chi tiet tung xa
        if detail:
            new_path = os.path.join(
                tk_path, ngaythang, "Thống kê chi tiết từng xã theo từng năm", huyen
            )
            Path(new_path).mkdir(parents=True, exist_ok=True)

            lst_ma_ndk = pd.read_sql(
                "select MaNoiDangKy from HT_NOIDANGKY where TenExport is not null and TenExport != '' ",
                conn,
            )["MaNoiDangKy"].to_list()

            for ma_ndk in lst_ma_ndk:
                ten_export = str(
                    pd.read_sql(
                        f"select TenExport from HT_NOIDANGKY where TenExport is not null and TenExport != '' and MaNoiDangKy = {ma_ndk}",
                        conn,
                    )["TenExport"].to_list()[0]
                )

                print(ten_export)

                file_truong = os.path.join(
                    new_path,
                    ten_export + " - truong_thong_tin.xlsx",
                )
                file_ban_ghi = os.path.join(
                    new_path,
                    ten_export + " - ban_ghi.xlsx",
                )

                tk_chi_tiet_theo_nam(
                    conn,
                    f"{ks} where noiDangKy = {ma_ndk}",
                    f"{kt} where noiDangKy = {ma_ndk}",
                    f"{kh} where noiDangKy = {ma_ndk}",
                    f"{cmc} where noiDangKy = {ma_ndk}",
                    f"{hn} where noiCap = {ma_ndk}",
                    file_ban_ghi,
                    file_truong,
                    ndk=ma_ndk,
                )

        # Chung 1 huyen
        if general:
            new_path = os.path.join(
                tk_path, ngaythang, "Thống kê huyện theo từng năm", huyen
            )
            Path(new_path).mkdir(parents=True, exist_ok=True)

            print(f"Tổng hợp huyện {huyen}")
            file_truong = os.path.join(
                new_path,
                f"Tổng hợp trường thông tin {huyen}.xlsx",
            )
            file_ban_ghi = os.path.join(
                new_path,
                f"Tổng hợp bản ghi {huyen}.xlsx",
            )
            tk_chi_tiet_theo_nam(
                conn, ks, kt, kh, cmc, hn, file_ban_ghi, file_truong, chi_tiet=False
            )


def pretty_excel(xlsx):
    workbook = openpyxl.load_workbook(xlsx)
    sheet = workbook["Thống kê hộ tịch"]

    new_data = [
        "",
        "15 ký tự trở xuống",
        "16 đến 50 ký tự",
        "Trên 50 ký tự",
        "15 ký tự trở xuống",
        "16 đến 50 ký tự",
        "Trên 50 ký tự",
        "15 ký tự trở xuống",
        "16 đến 50 ký tự",
        "Trên 50 ký tự",
        "15 ký tự trở xuống",
        "16 đến 50 ký tự",
        "Trên 50 ký tự",
        "15 ký tự trở xuống",
        "16 đến 50 ký tự",
        "Trên 50 ký tự",
    ]
    row_index = 2
    sheet.insert_rows(row_index)
    for col_num, value in enumerate(new_data, start=1):
        sheet.cell(row=row_index, column=col_num, value=value)

    merge_range = "B1:D1"
    sheet.merge_cells(merge_range)
    sheet["B1"] = "Khai sinh"

    merge_range = "E1:G1"
    sheet.merge_cells(merge_range)
    sheet["E1"] = "Khai tử"

    merge_range = "H1:J1"
    sheet.merge_cells(merge_range)
    sheet["H1"] = "Kết hôn"

    merge_range = "K1:M1"
    sheet.merge_cells(merge_range)
    sheet["K1"] = "Cha mẹ con"

    merge_range = "N1:P1"
    sheet.merge_cells(merge_range)
    sheet["N1"] = "Hôn nhân"

    merge_range = "A1:A2"
    sheet.merge_cells(merge_range)
    sheet["A1"] = "Năm"

    for cell in sheet[1]:
        cell.alignment = openpyxl.styles.Alignment(
            wrap_text=True, vertical="center", horizontal="center"
        )
        cell.font = openpyxl.styles.Font(bold=True)

    for cell in sheet[2]:
        cell.alignment = openpyxl.styles.Alignment(
            wrap_text=True, vertical="center", horizontal="center"
        )
        cell.font = openpyxl.styles.Font(bold=True)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    workbook.save(xlsx)


# while True:
#     if datetime.datetime.now().strftime("%H") in (
#         "12",
#         "21",
#     ) and datetime.datetime.now().strftime("%M") in ("00"):
#         tk_chi_tiet_theo_nam_xa(
#             fmt=True,
#             status=True,
#             book=True,
#             year=True,
#             period=False,
#             detail=True,
#             general=True,
#         )

#         print(datetime.datetime.now())


tk_chi_tiet_theo_nam_xa(
    fmt=True,
    status=True,
    book=True,
    year=True,
    period=False,
    detail=True,
    general=True,
)
