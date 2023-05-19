import os
import re

from bs4 import BeautifulSoup
from openpyxl import Workbook

for root, dirs, files in os.walk(r'C:\M-Files Content Packages\Kho của tôi'):
    for file in files:
        pattern = re.compile(".*Content1.xml$")

        if pattern.match(file):

            with open(os.path.join(root, file), 'r', encoding="utf-8") as f:
                data = f.read()

            Bs_Data_All = BeautifulSoup(data, "xml")

            dict = {}
            dict['Đường dẫn'] = []
            
            for Bs_data in Bs_Data_All.find_all('object'):
                # properties = Bs_data.find_all('property', {'name': 'Tên hoặc Tiêu đề'})
                property = Bs_data.find_all('property')
                docfile = Bs_data.find_all('docfile')
                # docfile = Bs_data.find_all('docfile')[0]['pathfrombase']
                properties = Bs_data.find_all('properties')

                workbook = Workbook()
                sheet = workbook.active
                sheet.title = 'data'
                # sheet.column_dimensions['A'].width = 50
                # sheet.column_dimensions['B'].width = 100

                # append to dict
                for pro in property:
                    # print("{0:50}\t{1}".format(property['name'], property.text))
                    if(pro['name'] not in ('Ngày tạo', 'Người tạo', 'Kích thước trên máy chủ (tất cả các bản)', 
                                           'Đánh dấu lưu trữ', 'Đường dẫn ban đầu (1/3)', 'Đổi tượng đã thay đổi', 
                                           'Sửa đổi gần nhất', 'Một tệp', 'Người sửa gần nhất', 'Thay đổi trạng thái', 
                                           'Kích thước trên máy chủ (bản này)')):
                        if pro['name'] not in dict:
                            dict[pro['name']] = [pro.text]
                        else:
                            dict[pro['name']].append(pro.text)

                dict['Đường dẫn'].append(Bs_data.find('docfile')['pathfrombase'])

                col = 1
                row = 1
                for key, values in dict.items():
                    for value in values:
                        row += 1
                        # print title
                        sheet.cell(column=col, row=1, value="{0}".format(key))

                        # print value
                        sheet.cell(column=col, row=row,
                                   value="{0}".format(value))
                    row = 1
                    col += 1

            # save excel
            workbook.save(filename=r'C:\Users\Nam\Downloads\M-Files\data.xlsx')
            workbook.close()
