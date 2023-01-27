import collections
import configparser
import datetime
import glob
import itertools
import json
import os
import pathlib
import random
import re
import shutil
import string
import time
import urllib.parse
from collections import Counter
from functools import wraps
from itertools import cycle
from pathlib import Path
from sys import stdout as terminal
from threading import Thread
from time import sleep

import img2pdf
import inquirer
import numpy as np
import openpyxl
import pandas as pd
import pikepdf
import pyautogui
import PyPDF2
import pyperclip
import xlsxwriter
from getpass4 import getpass
from openpyxl import Workbook
from pdf2image import convert_from_path
from PIL import Image
from PyPDF2 import PdfFileMerger, PdfFileReader, PdfFileWriter

dictEx = {}
totalPdfPages = 0
countDirs = 0
countFiles = 0
countFilesNotThumbs = 0
arr = np.array([])
countModifier = 0
folderPath1 = ''
folderPath2 = ''
filePath = ''
A4_SIZE = 8699840


# config_file = os.path.join(os.getenv('APPDATA'), 'filex.config')

chucnang = {"Phân tích thư mục": "Y",
            "Phân tích file txt": "Y",
            "Tìm kiếm tệp tin": "Y",
            "Kiểm tra trùng tên": "Y",
            "Thời gian tệp tin thay đổi": "Y",
            "Tạo cấu trúc theo tên tệp tin": "Y",
            "Sửa đổi metadate pdf": "Y",
            "JPG => PDF": "Y",
            "PDF => JPG": "Y",

            "Cấu hình": "Y"}


# decor
def get_files(function):
    @wraps(function)
    def wrapper(folderPath, fileFormat=''):
        global countDirs
        global countFiles
        global countFilesNotThumbs

        for root, dirs, files in os.walk(folderPath):
            for file in files:
                # $ regex kết thúc là ký tự trc $
                pattern = re.compile(".*"+fileFormat+"$")
                if pattern.match(file):
                    function(root, file)
            countDirs += len(dirs)
            countFiles += len(files)
    return wrapper


# animation loading cmd
def loading(function):
    @wraps(function)
    def wrapper():
        done = False

        def clear():
            return os.system('cls')

        def animate():
            for c in itertools.cycle(['⠷', '⠯', '⠟', '⠻', '⠽', '⠾']):
                if done:
                    break
                terminal.write('\rLoading ' + c + ' ')
                terminal.flush()
                sleep(0.1)
            # terminal.write('\rDone!    ')
            terminal.flush()

        t = Thread(target=animate)
        t.start()

        # Chay lenh tai day
        function()

        # clear()
        done = True

    return wrapper


# sửa đổi metadata file pdf
@get_files
def edit_metadata_pdf(root, file):
    try:
        pdf = pikepdf.open(os.path.join(root, file))
        with pdf.open_metadata() as meta:
            # meta['dc:title'] = ""
            print(meta['dc:title'])
    except Exception as e:
        print(str(e))


# phân tích thư mục
@loading
def analysis_in_folder():

    @get_files
    def inner(root, file):
        global dictEx
        global totalPdfPages

        # thống kê số file của từng loại file
        if (os.path.splitext(file)[1] in dictEx.keys()):
            dictEx[os.path.splitext(file)[1]] += 1
        else:
            dictEx[os.path.splitext(file)[1]] = 1

        # đếm trang pdf
        if ('.pdf' in os.path.splitext(file)[1]):
            readpdf = PyPDF2.PdfFileReader(open(os.path.join(root, file), 'rb'), strict=False)
            totalPdfPages += readpdf.numPages

    # không cần gán tham số fileFomat vào wapper bởi đã có giá trị mặc định
    inner(folderPath1)

    print(f"\rTổng số thư mục: \t{countDirs: >5,}\nTổng số file: \t\t{countFiles: >5,} \n")
    for key, value in dictEx.items():
        print(f"\rSố file {key: <22} {value}")

    print(f'\n\rSố trang pdf: {totalPdfPages:>18}\n')


# lấy tất cả file
@get_files
def get_fullname(root, file):
    global arr
    arr = np.append(arr, os.path.join(root, file))


# tìm kiếm trùng tên nhiều thư mục
@get_files
def get_basename(root, file):
    global arr
    arr = np.append(arr, file)


@get_files
def check_modifier_file(root, file):
    global countModifier
    date_modifier = os.path.getmtime(os.path.join(root, file))
    date_create = os.path.getctime(os.path.join(root, file))
    getDate = datetime.datetime.now().strftime("%d/%m/%Y")
    if (datetime.date.fromtimestamp(max(date_modifier, date_create)).strftime("%d/%m/%Y") in getDate):
        countModifier += 1


# Tạo cấu trúc thư mục như Files
@get_files
def create_struct(root, file):
    try:
        head, tail = (os.path.split(
            Path(os.path.join(root, file))))
        if (len(tail.split('.')) >= 6):
            newPath = os.path.join(pathTarget, tail.split('.')[0], tail.split('.')[
                1], tail.split('.')[2], tail.split('.')[3])
            Path(newPath).mkdir(parents=True, exist_ok=True)
            if not os.path.exists(os.path.join(newPath, tail.replace(' ', ''))):
                shutil.move(os.path.join(root, file), os.path.join(
                    newPath, tail.replace(' ', '')))
            else:
                pass
    except:
        pass


def config():
    global chucnang

    questions = [
        inquirer.Checkbox(
            "config",
            message="Chọn các chức năng?",
            choices={k: v for (k, v) in chucnang.items()}.keys(),
            default={k: v for (k, v) in chucnang.items() if 'Y' in v}.keys(),
        ),
    ]

    # class WorkplaceFriendlyTheme(Default):
    #     """Custom theme replacing X with Y and o with N"""

    # def __init__(self):
    #     super().__init__()
    #     self.Checkbox.selected_icon = "Y"
    #     self.Checkbox.unselected_icon = "N"

    # answers = inquirer.prompt(questions, theme=WorkplaceFriendlyTheme())
    answers = inquirer.prompt(questions)

    keychucnang = chucnang.keys()
    for item in keychucnang:
        if (item in answers['config']):
            chucnang[item] = 'Y'
        else:
            chucnang[item] = 'N'

    chucnang['Cấu hình'] = 'Y'

    # with open(config_file, mode='w', encoding="utf8") as f:
    #     f.write(str(chucnang))


def hidden_password():
    password = getpass('Password: ')
    print(password)


def jpg2pdf(img_path, pdf_path):

    # opening image
    image = Image.open(img_path)

    # converting into chunks using img2pdf
    pdf_bytes = img2pdf.convert(image.filename)

    # opening or creating pdf file
    file = open(pdf_path, "wb")

    # writing pdf files with chunks
    file.write(pdf_bytes)

    # closing image file
    image.close()

    # closing pdf file
    file.close()


def pdf2jpg(pdfPath, jpgPath):
    # pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    Image.MAX_IMAGE_PIXELS = 1000000000000

    for root, dirs, files in os.walk(pdfPath):
        for file in files:

            pattern = re.compile(".*pdf$")
            if pattern.match(file):
                pages = convert_from_path(
                    os.path.join(root, file), 500, poppler_path=r'library\poppler-22.12.0\Library\bin')
                image_counter = 1
                for page in pages:
                    filename = os.path.splitext(file)[0] + '#' + str(image_counter).zfill(5)+".jpg"
                    page.save(os.path.join(jpgPath, filename), 'JPEG')
                    image_counter = image_counter + 1


def textfile_analysis(textfile):
    # script_name = sys.argv[0]
    script_name = 'filex.py'

    res = {
        "total_lines": "",
        "total_characters": "",
        "total_words": "",
        "unique_words": "",
        "special_characters": ""
    }

    try:
        # textfile = sys.argv[1]
        with open(textfile, "r", encoding="utf_8") as f:

            data = f.read()
            res["total_lines"] = data.count(os.linesep)
            res["total_characters"] = len(data.replace(
                " ", "")) - res["total_lines"]
            counter = collections.Counter(data.split())
            d = counter.most_common()
            res["total_words"] = sum([i[1] for i in d])
            res["unique_words"] = len([i[0] for i in d])
            special_chars = string.punctuation
            res["special_characters"] = sum(v for k, v in collections.Counter(
                data).items() if k in special_chars)

    except IndexError:
        print(f'\rUsage: {script_name} TEXTFILE')
    except IOError:
        print(f'\r{textfile} cannot be opened.')

    print(f'\r{res}')


def format_str(strA):
    specialcharacters = ['\\', '/', ':', '*', '?', '"', ',', '<', '>', '|']
    for i in strA:
        if (i in specialcharacters):
            strA = strA.replace(i, "")

    return str(strA)


def compound_unicode(unicode_str):
    """
    Chuyển đổi chuỗi Unicode Tổ Hợp sang Unicode Dựng Sẵn
    """
    dict_compound = {'\u0065\u0309': '\u1EBB', '\u0065\u0301': '\u00E9', '\u0065\u0300': '\u00E8',
                     '\u0065\u0323': '\u1EB9', '\u0065\u0303': '\u1EBD', '\u00EA\u0309': '\u1EC3',
                     '\u00EA\u0301': '\u1EBF', '\u00EA\u0300': '\u1EC1', '\u00EA\u0323': '\u1EC7',
                     '\u00EA\u0303': '\u1EC5', '\u0079\u0309': '\u1EF7', '\u0079\u0301': '\u00FD',
                     '\u0079\u0300': '\u1EF3', '\u0079\u0323': '\u1EF5', '\u0079\u0303': '\u1EF9',
                     '\u0075\u0309': '\u1EE7', '\u0075\u0301': '\u00FA', '\u0075\u0300': '\u00F9',
                     '\u0075\u0323': '\u1EE5', '\u0075\u0303': '\u0169', '\u01B0\u0309': '\u1EED',
                     '\u01B0\u0301': '\u1EE9', '\u01B0\u0300': '\u1EEB', '\u01B0\u0323': '\u1EF1',
                     '\u01B0\u0303': '\u1EEF', '\u0069\u0309': '\u1EC9', '\u0069\u0301': '\u00ED',
                     '\u0069\u0300': '\u00EC', '\u0069\u0323': '\u1ECB', '\u0069\u0303': '\u0129',
                     '\u006F\u0309': '\u1ECF', '\u006F\u0301': '\u00F3', '\u006F\u0300': '\u00F2',
                     '\u006F\u0323': '\u1ECD', '\u006F\u0303': '\u00F5', '\u01A1\u0309': '\u1EDF',
                     '\u01A1\u0301': '\u1EDB', '\u01A1\u0300': '\u1EDD', '\u01A1\u0323': '\u1EE3',
                     '\u01A1\u0303': '\u1EE1', '\u00F4\u0309': '\u1ED5', '\u00F4\u0301': '\u1ED1',
                     '\u00F4\u0300': '\u1ED3', '\u00F4\u0323': '\u1ED9', '\u00F4\u0303': '\u1ED7',
                     '\u0061\u0309': '\u1EA3', '\u0061\u0301': '\u00E1', '\u0061\u0300': '\u00E0',
                     '\u0061\u0323': '\u1EA1', '\u0061\u0303': '\u00E3', '\u0103\u0309': '\u1EB3',
                     '\u0103\u0301': '\u1EAF', '\u0103\u0300': '\u1EB1', '\u0103\u0323': '\u1EB7',
                     '\u0103\u0303': '\u1EB5', '\u00E2\u0309': '\u1EA9', '\u00E2\u0301': '\u1EA5',
                     '\u00E2\u0300': '\u1EA7', '\u00E2\u0323': '\u1EAD', '\u00E2\u0303': '\u1EAB',
                     '\u0045\u0309': '\u1EBA', '\u0045\u0301': '\u00C9', '\u0045\u0300': '\u00C8',
                     '\u0045\u0323': '\u1EB8', '\u0045\u0303': '\u1EBC', '\u00CA\u0309': '\u1EC2',
                     '\u00CA\u0301': '\u1EBE', '\u00CA\u0300': '\u1EC0', '\u00CA\u0323': '\u1EC6',
                     '\u00CA\u0303': '\u1EC4', '\u0059\u0309': '\u1EF6', '\u0059\u0301': '\u00DD',
                     '\u0059\u0300': '\u1EF2', '\u0059\u0323': '\u1EF4', '\u0059\u0303': '\u1EF8',
                     '\u0055\u0309': '\u1EE6', '\u0055\u0301': '\u00DA', '\u0055\u0300': '\u00D9',
                     '\u0055\u0323': '\u1EE4', '\u0055\u0303': '\u0168', '\u01AF\u0309': '\u1EEC',
                     '\u01AF\u0301': '\u1EE8', '\u01AF\u0300': '\u1EEA', '\u01AF\u0323': '\u1EF0',
                     '\u01AF\u0303': '\u1EEE', '\u0049\u0309': '\u1EC8', '\u0049\u0301': '\u00CD',
                     '\u0049\u0300': '\u00CC', '\u0049\u0323': '\u1ECA', '\u0049\u0303': '\u0128',
                     '\u004F\u0309': '\u1ECE', '\u004F\u0301': '\u00D3', '\u004F\u0300': '\u00D2',
                     '\u004F\u0323': '\u1ECC', '\u004F\u0303': '\u00D5', '\u01A0\u0309': '\u1EDE',
                     '\u01A0\u0301': '\u1EDA', '\u01A0\u0300': '\u1EDC', '\u01A0\u0323': '\u1EE2',
                     '\u01A0\u0303': '\u1EE0', '\u00D4\u0309': '\u1ED4', '\u00D4\u0301': '\u1ED0',
                     '\u00D4\u0300': '\u1ED2', '\u00D4\u0323': '\u1ED8', '\u00D4\u0303': '\u1ED6',
                     '\u0041\u0309': '\u1EA2', '\u0041\u0301': '\u00C1', '\u0041\u0300': '\u00C0',
                     '\u0041\u0323': '\u1EA0', '\u0041\u0303': '\u00C3', '\u0102\u0309': '\u1EB2',
                     '\u0102\u0301': '\u1EAE', '\u0102\u0300': '\u1EB0', '\u0102\u0323': '\u1EB6',
                     '\u0102\u0303': '\u1EB4', '\u00C2\u0309': '\u1EA8', '\u00C2\u0301': '\u1EA4',
                     '\u00C2\u0300': '\u1EA6', '\u00C2\u0323': '\u1EAC', '\u00C2\u0303': '\u1EAA'}

    for k, v in dict_compound:
        unicode_str = unicode_str.replace(k, v)
    return unicode_str


def string_search_from_multiple_files(path):
    text = input("input text : ")

    # path = input("path : ")
    f = 0
    os.chdir(path)
    files = os.listdir()
    # print(files)
    for file_name in files:
        abs_path = os.path.abspath(file_name)
        if os.path.isdir(abs_path):
            string_search_from_multiple_files(abs_path)
        if os.path.isfile(abs_path):
            f = open(file_name, "r")
            if text in f.read():
                f = 1
                print(text + " found in ")
                final_path = os.path.abspath(file_name)
                print(final_path)
                return True

    if f == 1:
        print(text + " not found! ")
        return False


def removeEscape(value):
    return ' '.join(str(value).splitlines()).strip()


def split_pdf(pathPdfInput, pathPdfOutput):
    for root, dirs, files in os.walk(pathPdfInput):
        for file in files:
            pattern = re.compile(".*pdf$")

            if pattern.match(file):
                inputpdf = PdfFileReader(os.path.join(root, file), strict=False)
                fileName = pathlib.Path(file).stem
                parentPath = pathlib.Path(file).parent.absolute()

                pdfs = []

                for i in range(inputpdf.numPages):
                    output = PdfFileWriter()
                    output.addPage(inputpdf.getPage(i))
                    with open("%s#%s.pdf" % (os.path.join(pathPdfOutput, fileName), str(i).zfill(5)), "wb") as outputStream:
                        output.write(outputStream)
                        pdfs.append("%s#%s.pdf" % (os.path.join(pathPdfOutput, fileName), i))
                        outputStream.close()


def merge_pdf(lstFilesInput, fileNameOutput):
    merger = PdfFileMerger(strict=False)
    for file in lstFilesInput:
        merger.append(file)
    merger.write(fileNameOutput)


def setDpiImg2Pdf():
    # storing image path
    img_path = r"C:\Users\Nam\Downloads\New folder\result.jpg"

    # storing pdf path
    pdf_path = r"C:\Users\Nam\Downloads\New folder\result.pdf"

    dpix = dpiy = 300
    layout_fun = img2pdf.get_fixed_dpi_layout_fun((dpix, dpiy))

    # opening image
    image = Image.open(img_path)

    # converting into chunks using img2pdf
    pdf_bytes = img2pdf.convert(image.filename, layout_fun=layout_fun)

    # opening or creating pdf file
    file = open(pdf_path, "wb")

    # writing pdf files with chunks
    file.write(pdf_bytes)

    # closing image file
    image.close()

    # closing pdf file
    file.close()

    # output
    print("Successfully made pdf file")


def detect_size(img):
    # Disable PIL DecompositionBomb threshold for reading large images.
    pil_max_px = Image.MAX_IMAGE_PIXELS
    Image.MAX_IMAGE_PIXELS = None
    im = Image.open(img)
    Image.MAX_IMAGE_PIXELS = pil_max_px

    print('{}'.format(im.size))
    print((im.width*im.height)/A4_SIZE)


def spit_and_merge_pdf(pathPdfInput, bytes=10485760):
    if (os.path.getsize(pathPdfInput) >= 10485760):
        inputpdf = PdfFileReader(pathPdfInput, "rb")
        fileName = pathlib.Path(pathPdfInput).stem
        parentPath = pathlib.Path(pathPdfInput).parent.absolute()

        sum = 0
        pdfs = []

        for i in range(inputpdf.numPages):
            output = PdfFileWriter()
            output.addPage(inputpdf.getPage(i))
            with open("%s.%s.pdf" % (fileName, i), "wb") as outputStream:
                output.write(outputStream)
                pdfs.append("%s.%s.pdf" % (fileName, i))
                outputStream.close()

        merger = PdfFileMerger()
        index = 0
        for page in pdfs:
            # get size file
            sum += os.path.getsize(page)
            # nếu dung lượng các trang chưa quá 10MB thì vẫn thêm vào sau
            if (sum < bytes):
                merger.append(page)

            else:
                if (page == pdfs[-1]):
                    merger.append(page)
                merger.write(r"%s\%s.%s.pdf" %
                             (parentPath, fileName, index + 1))
                with open(r"split.txt", "a", encoding="utf-8") as fp:
                    fp.write(r"%s\%s.%s.pdf" %
                             (parentPath, fileName, index + 1) + '\n')
                merger = PdfFileMerger()
                merger.append(page)
                index += 1
                sum = os.path.getsize(page)

        try:
            listPdf = glob.glob("*pdf")
            for page in listPdf:
                os.remove(page)
                # page.unlink()
        except:
            pass


def split_merge_pdf_ocr(folderPath):
    # lấy danh sách tên file cần ghép
    for root, dirs, files in os.walk(folderPath):
        dict = {}
        for file in files:
            if ('#' in file):

                i = os.path.join(root, file.split('#')[0] + '.pdf')

                if (i not in dict.keys()):
                    dict[i] = []

                dict[i].append(os.path.join(root, file))

    # thực hiện ghép
    for k, v in dict.items():
        merge_pdf(v, k)


def write_to_excel(arr, filename):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'data'
    for row in range(1, len(arr)+1):
        for col in range(1, len(arr[0])+1):
            sheet.cell(column=col, row=row,
                       value="{0}".format(arr[row-1][col-1]))

    workbook.save(filename=filename)
    workbook.close()


def read_from_excel(path):
    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)

    # Get workbook active sheet object
    # from the active attribute
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row

    # list all row excel
    lst = []

    # list title
    lstTitle = []

    # Loop will print all rows name
    for i in range(1, max_row + 1):
        # list noi dung
        arr = []

        # Loop will print all column name
        for j in range(1, max_col + 1):
            # doc thong tin trong o
            cell_obj = sheet_obj.cell(row=i, column=j)

            if (cell_obj.value is not None):
                # compound unicode cell_obj
                content = compound_unicode(str(cell_obj.value))

                # nếu là dòng đầu tiên thì thêm vào list tiêu đề
                if (i == 1):
                    lstTitle.append(content)
                else:
                    arr.append(content)
            else:
                # nếu trống thông tin list noi dung se them ''
                arr.append('')

        # thêm vào danh sách nội dung dòng
        lst.append(arr)

    return lst


def change_column():
    # use glob to get all the csv files
    # in the folder
    path = r'C:\Users\ADMIN\Downloads\New folder (2)'
    csv_files = glob.glob(os.path.join(path, "*.xlsx"))

    # loop over the list of excel files
    for f in csv_files:

        # read the excel file
        df = pd.read_excel(f)

        df_new = pd.read_excel(f)

        # set header
        df_new.columns = ['Số tt', 'Sổ KHVB', 'Ngày tháng VB',
                          'Trích yếu nội dung', 'Tác giả văn bản', 'Tờ số', 'ghi chú']

        df_new['Số tt'] = df['Số tt']
        df_new['Sổ KHVB'] = df['Sổ KHVB']
        df_new['Ngày tháng VB'] = df['Ngày tháng VB']
        df_new['Trích yếu nội dung'] = df['Trích yếu nội dung']
        df_new['Tác giả văn bản'] = df['Tác giả văn bản']
        df_new['Tờ số'] = df['Tờ số']
        df_new['ghi chú'] = df['ghi chú']

        os.remove(f)
        writer = pd.ExcelWriter(f, engine='xlsxwriter')

        # df_new.to_excel(f, index=False)

        df_new.to_excel(writer, sheet_name='Sheet1', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        border_fmt = workbook.add_format(
            {'bottom': 2, 'top': 2, 'left': 2, 'right': 2})
        worksheet.conditional_format(xlsxwriter.utility.xl_range(0, 0, len(df_new), len(
            df_new.columns) - 1), {'type': 'no_errors', 'format': border_fmt})

        format_header = workbook.add_format()
        format_header.set_valign('vcenter')
        format_header.set_align('center')
        format_header.set_bold()
        format_header.set_text_wrap()

        format_data = workbook.add_format()
        format_data.set_valign('vcenter')
        format_data.set_align('center')
        format_data.set_text_wrap()

        worksheet.set_column('A:Z', 25, format_data)
        worksheet.set_row(0, 30, format_header)

        writer.save()


def spam():
    print("Tool Spam 1.0")
    msg = input("Nhập nội dung cần spam: ").split(" || ")
    n = int(input("Nhập số lần Spam: "))
    m = float(input("Nhập thời gian delay: "))

    print("Chuẩn bị")
    # Đếm ngược 5 giây
    for i in range(5, 0, -1):
        print(i, end="...", flush='True')
        time.sleep(1)
    print("Bắt đầu")

    # SPAM
    for i in range(n):
        pyperclip.copy(random.choice(msg))
        pyautogui.hotkey("ctrl", "v")
        pyautogui.press("enter")
        time.sleep(m)  # Delay


def auto_write():
    # while True:
    # sohieu = input('Số hiệu: ')
    # name = input('Tên: ')
    # date = input('Ngày: ')

    # # 432,250
    # # 83, 730

    # # điền số hiệu
    # pyperclip.copy(sohieu + ' /STP-LLTP')
    # pyautogui.click(432,250)
    # pyautogui.hotkey("ctrl", "v")

    # # điền trích yếu
    # pyperclip.copy('PHIẾU LÝ LỊCH TƯ PHÁP CỦA ' + name)
    # pyautogui.press('tab')
    # pyautogui.hotkey("ctrl", "v")

    # # điền ngày tháng có hiệu lực
    # pyperclip.copy(date)
    # pyautogui.press('tab')
    # pyautogui.hotkey("ctrl", "v")

    # # điền ngày tháng ban hành
    # pyperclip.copy(date)
    # pyautogui.press('tab')
    # pyautogui.press('tab')
    # pyautogui.hotkey("ctrl", "v")

    # # điền tên tổ chức
    # pyperclip.copy(name)
    # pyautogui.press('tab')
    # pyautogui.press('tab')
    # pyautogui.hotkey("ctrl", "v")

    # pyautogui.click(83, 730)
    # pyautogui.hotkey("alt", "tab")
    # print('------------------------------')

    # crawl data website => chuyển thành mp3

    # pyautogui.moveTo(83, 730)

    pyautogui.hotkey("alt", "tab")
    for i in range(318):
        pyautogui.press('F2')
        pyautogui.press('right')
        pyautogui.write('(1)')
        pyautogui.press('tab')


def auto_click():
    # print ('di chuyển chuột đến vị trí cần click')
    # time.sleep(5)

    # x, y = pyautogui.position()

    # i, j = pyautogui.position()

    # for _ in range(10):
    #     pyautogui.click(x, y)
    #     time.sleep(1)

    # pyautogui.press('a')

    for _ in range(10000000):
        # Point(x=1809, y=131): tải xuống
        # Point(x=1594, y=635): save
        # time.sleep(1)
        pyautogui.click(1291, 595)
        # time.sleep(1)
        # pyautogui.click(1594, 635)
        # pyautogui.hotkey('ctrl', 'w')

        # pyautogui.hotkey('ctrl', 'c')
        print(pyautogui.position())


if __name__ == '__main__':

    # if not (
    #         os.path.exists(config_file)):
    #     with open(config_file, mode='a', encoding="utf8") as f:
    #         f.write(str(chucnang))

    # else:
    #     if os.path.getsize(config_file) == 0:
    #         os.remove(config_file)

    #         with open(config_file, mode='w', encoding="utf8") as f:
    #             f.write(str(chucnang))
    #     else:
    #         with open(config_file, mode='r', encoding="utf8") as f:
    #             chucnang = json.loads(f.readlines()[0].replace('\'', '\"'))

    while True:
        print("-----Công cụ làm việc với Files-----")
        print()
        arr = np.array([])

        try:
            questions = [
                inquirer.List(
                    "choise", message="Chức năng cần sử dụng?",
                    # lọc những chức năng cho phép
                    choices={k: v for (k, v) in chucnang.items() if 'Y' in v}.keys())]

            answers = inquirer.prompt(questions)

            match answers['choise']:
                case 'Phân tích thư mục':
                    folderPath1 = input('Nhập đường dẫn cần phân tích: ')
                    analysis_in_folder()
                    
                case 'Phân tích file txt':
                    txtPath = input('Nhập đường đẫn file text: ')

                    @loading
                    def run():
                        textfile_analysis(txtPath)
                    run()

                case 'Tìm kiếm tệp tin':
                    fileSearch = input('Nhập tên file: ')
                    get_fullname(input('Nhập đường dẫn cần tìm kiếm: '))
                    for file in arr:
                        if (fileSearch in file):
                            print(file)

                case 'Kiểm tra trùng tên':
                    for path in input('Nhập đường dẫn cần kiểm tra trùng: ').split(','):
                        get_basename(path)

                    dup = {item for item, count in Counter(arr).items() if count > 1}
                    print(len(dup))
                    print(dup)

                case 'Thời gian tệp tin thay đổi':
                    check_modifier_file(input('Nhập đường dẫn cần kiểm tra'))
                    print(countModifier)

                case 'Tạo cấu trúc theo tên tệp tin':
                    pathTarget = input('Nhập thư mục đầu ra: ')
                    create_struct(input('Nhập thư mục đầu vào: '))

                case 'Sửa đổi metadate pdf':
                    edit_metadata_pdf(input('Nhập đường dẫn cần sửa dổi: '))

                case 'JPG => PDF':
                    dir_img = input('Nhập đường dẫn thư mục jpg: ')
                    dir_pdf = input('Nhập đường dẫn thư mục pdf: ')

                    for root, dirs, files in os.walk(dir_img):
                        for file in files:
                            try:
                                jpg2pdf(os.path.join(root, file),
                                        os.path.join(dir_pdf, os.path.splitext(file)[0] + '.pdf'))
                            except Exception as e:
                                pass

                case 'PDF => JPG':
                    pdfPath = input('Nhập đường dẫn thư mục pdf: ')
                    jpgPath = input('Nhập đường dẫn thư mục jpg: ')

                    @loading
                    def run():
                        pdf2jpg(pdfPath, jpgPath)

                    run()

                case 'Cấu hình':
                    config()

                case _:
                    pass

        except Exception as e:
            print(str(e))
            input()
        finally:
            yn = str(input('\rBạn có muốn tiếp tục? (Y/N?): ')).upper()
            if (yn == 'Y'):
                os.system('cls')
                continue
            else:
                print("Hẹn gặp lại!!!\n")
                break

# split_pdf(r'E:\tay ninh\ubnd 2014 chua nen', r'E:\tay ninh\ubnd 2014 chua nen')

# split_merge_pdf_ocr(r'E:\tay ninh\da ocr\tay ninh\ubnd 2014 chua nen\tach')
